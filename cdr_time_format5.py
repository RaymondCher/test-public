#!/usr/bin/python
# -*- coding: UTF-8 -*-

# Purpose: 详单时间格式转化以及，用于对OP运营商管理系统导出来的详单文件时间格式进行处理
# V1.0 2021-03-13， Author: Xiao Tongquan
# V1.1 2022-06-28   发给客户的简化版文件增加剔除tj.gxapn功能，

import os
import time
import datetime
import win32com.client  #首先安装 pip install pywin32
#import win32com.client as win32    #首先安装 pip install pywin32
import openpyxl
from openpyxl.styles import PatternFill, colors, Alignment
import shutil
import pandas as pd
from pandas import Series,DataFrame

from openpyxl import load_workbook

# excel文件xls格式转换成xlsx格式
def xls2xlsx(fileName):
    #excel = win32.gencache.EnsureDispatch('Excel.Application')
    excel = win32com.client.DispatchEx('Excel.Application')
    wb = excel.Workbooks.Open(fileName)
    #wb.SaveAs()
    wb.SaveAs(fileName + "x", FileFormat=51)  # FileFormat = 51 is for .xlsx extension
    fileName = fileName + "x"
    wb.Close()                                # FileFormat = 56 is for .xls extension
    excel.Application.Quit()                  # 后缀名的大小写不通配，需按实际修改：xls，或XLS
    return(fileName)

# 将绝对毫秒转换成 yyyy-mm-dd hh:mm:ss格式
def milisecond_convert(timeStamp):
    timeStamp = timeStamp
    timeStamp /= 1000.0
    timeArray = time.localtime(timeStamp)
    otherStyleTime = time.strftime("%Y-%m-%d %H:%M:%S", timeArray)
    return(otherStyleTime)

def delete_file(file_name):
    if os.path.exists(file_name):
        os.remove(file_name)
    else:
        print('要删除的临时文件不存在!!!')
        pass
    # C:\Users\hp\AppData\Local\Temp\gen_py\3.6\00020813-0000-0000-C000-000000000046x0x1x9
    # 将yyyy/mm/ddhhmmss格式转换成yyyy-mm-dd hh:mm:ss格式

def yyyymmdd_convert(timeStamp):
    dt = datetime.datetime.strptime(timeStamp, '%Y%m%d%H%M%S')
    otherStyleTime = dt.strftime('%Y-%m-%d %H:%M:%S')
    return(otherStyleTime)

# 联通流量详单时间格式进行处理
def cuc_milisecond_time_final(fileName):

    # 读入文件并获得active的表单
    workbook = openpyxl.load_workbook(fileName)
    sheet = workbook.active

    # 插入第15,16两列，第1列从1开始数起
    sheet.insert_cols(idx=15, amount=2)
    sheet.cell(1, 15).value = '计费开始时间'
    sheet.cell(1, 16).value = '计费结束时间'

    fill_1 = PatternFill("solid", fgColor="BCEE68")     #对新增的两个单元格进行颜色填充
    sheet.cell(1, 15).fill = fill_1
    sheet.cell(1, 16).fill = fill_1

    workbook.save(fileName)

    # 重新读入文件，进行计费开始时间/结束时间 格式的转换.
    workbook = openpyxl.load_workbook(fileName)
    sheet = workbook.active

    # 原始文件的计费开始时间/结束时间分别是第M(13)列和第 N(14)列,在前面新插入的K（15）和列 L（16）列，用来储存转格式转换后的时间
    for row in range(2, sheet.max_row + 1):

        chargeStart = int(sheet.cell(row, 13).value)
        chargeEnd = int(sheet.cell(row, 14).value)

        chargeStartNew = milisecond_convert(chargeStart)
        chargeEndNew = milisecond_convert(chargeEnd)

        sheet.cell(row, 15).value = chargeStartNew
        sheet.cell(row, 16).value = chargeEndNew
    workbook.save(fileName)
    print('\n')
    #print('时间格式转换后的结果保存在：', fileName)

# 联通短信详单时间格式进行处理
def cuc_yyyymmdd_time_final(fileName):
    # 读入文件并获得active的表单
    workbook = openpyxl.load_workbook(fileName)
    sheet = workbook.active

    sheet.cell(1, 6).value = '计费开始时间'

    fill_1 = PatternFill("solid", fgColor="BCEE68")     #对新增的单元格进行颜色填充
    sheet.cell(1, 6).fill = fill_1

    workbook.save(fileName)

    # 重新读入文件，进行计费开始时间/结束时间 格式的转换.
    workbook = openpyxl.load_workbook(fileName)
    sheet = workbook.active

    # 原始文件的计费开始时间/结束时间分别是第E(5)列,使用第F(6)刘来储存转格式转换后的时间

    for row in range(2, sheet.max_row + 1):
        chargeStart = str(sheet.cell(row, 5).value)

        chargeStartNew = yyyymmdd_convert(chargeStart)

        sheet.cell(row, 6).value = chargeStartNew

    workbook.save(fileName)
    print('处理完毕，结果保存在：', fileName)

# 移动流量详单时间格式进行处理
def cmc_milisecond_time_final(fileName):
    # 读入文件并获得active的表单
    workbook = openpyxl.load_workbook(fileName)
    sheet = workbook.active

    # Insert two columns with titles and overwrite the old file.
    sheet.insert_cols(idx=4, amount=1)
    sheet.cell(1, 4).value = '计费结束时间'

    fill_1 = PatternFill("solid", fgColor="BCEE68")     #对新增的单元格进行颜色填充
    sheet.cell(1, 4).fill = fill_1

    workbook.save(fileName)

    # 重新读入文件，进行计费开始时间/结束时间 格式的转换.
    workbook = openpyxl.load_workbook(fileName)
    sheet = workbook.active

    # 原始文件的计费结束时间分别是第C(3)列,在前面新插入的D（4）列，用来储存转格式转换后的时间

    for row in range(2, sheet.max_row + 1):

        chargeEnd = int(sheet.cell(row, 3).value)
        chargeEndNew = milisecond_convert(chargeEnd)
        sheet.cell(row, 4).value = chargeEndNew

    workbook.save(fileName)

    print('处理完毕，结果保存在：', fileName)

# 移动短信详单时间格式进行处理
def cmc_yyyymmdd_time_final(fileName):
    # 读入文件并获得active的表单
    workbook = openpyxl.load_workbook(fileName)
    sheet = workbook.active

    # Insert two columns with titles and overwrite the old file.
    sheet.insert_cols(idx=7, amount=2)     # 在第7列“finalState”前面插入2列
    sheet.cell(1, 7).value = '计费开始时间'
    sheet.cell(1, 8).value = '计费结束时间'

    fill_1 = PatternFill("solid", fgColor="BCEE68")     #对新增的两个单元格进行颜色填充
    sheet.cell(1, 7).fill = fill_1
    sheet.cell(1, 8).fill = fill_1
    workbook.save(fileName)

    # 重新读入文件，进行计费开始时间/结束时间 格式的转换.
    workbook = openpyxl.load_workbook(fileName)
    sheet = workbook.active

    # 原始文件的计费开始时间/结束时间分别是第E(5)列和第 F(6)列,在前面新插入的G（7）列和 H（8）列，用来储存转格式转换后的时间

    for row in range(2, sheet.max_row + 1):
        chargeStart = str(sheet.cell(row, 5).value)
        chargeEnd = str(sheet.cell(row, 6).value)

        chargeStartNew = yyyymmdd_convert(chargeStart)
        chargeEndNew = yyyymmdd_convert(chargeEnd)

        sheet.cell(row, 7).value = chargeStartNew
        sheet.cell(row, 8).value = chargeEndNew

    workbook.save(fileName)
    print('处理完毕，结果保存在：', fileName)


# 定义函数，对excel正文内容做居中处理
def excel_center(file_name):

    alignment = Alignment(horizontal='center',
                          vertical='center',
                          text_rotation=0,
                          indent=0)

    workbook = load_workbook(file_name)
    sheet = workbook.active
    rows = sheet.max_row        # ws.max_row 计算工作表行数，
    columns = sheet.max_column  # ws.max_column 计算作表列数

    for i in range(1, rows+1):  #从第1行开始
        for j in range(1, columns+1):  #从第1列开始
            sheet.cell(i, j).alignment = alignment
    workbook.save(file_name)

# 处理联通理流量data文件
def cuc_data_process():
    # 因为python版本问题，下面这个临时文件夹需要提前删除掉，避免程序报下面的错误
    # AttributeError: module 'win32com.gen_py.00020813-0000-0000-C000-000000000046x0x1x9' has no attribute
    'CLSIDToClassMap'
    '''
    寻找下面报错要删除文件的路径
    import sys
    from win32com.client.gencache import EnsureDispatch
    xl = EnsureDispatch("Word.Application")
    print(sys.modules.[xl._module_]._file_)
    上面4条指令如果找不到的话，可以在windows上用everything 搜索gen_py
    '''
    #提前删除python运行过程中产生的一个临时文件，避免后面报错
    #file_to_bo_deleted = 'C:\\Users\\xiaotongquan\\AppData\\Local\\Temp\\gen_py\\3.6\\00020813-0000-0000-C000-000000000046x0x1x9'  # 文件夹路径
    #path_dir = 'C:\\Users\\xiaotongquan\\AppData\\Local\\Temp\\gen_py\\3.6' # 要删除的文件夹路劲
    #shutil.rmtree(path)

    # if os.path.exists(file_to_bo_deleted):
    #     #os.remove(file_to_bo_deleted)
    #     shutil.rmtree(path_dir)
    # else:
    #     pass

    global raw_data_after_time_formated
    PATH = 'C:\\Users\\raymondcher\\Downloads\\'  # PATH 定义了数据文件的绝对路径
    print('本次要处理的文件储存路径为：%s' % PATH)
    print()
    fileName = input('请输入计费详单文件名（不需要带后缀.xls)： 若直接回车,则默认输入文件名为 "联通流量详单.xls" : ')

    if fileName == "":
        raw_file_after_time_formated = PATH + '1_时间处理后的原始文件_联通流量详单.xlsx'  # 时间处理后的原始文件
        raw_file_after_time_formated_bk = PATH + '备份_1_时间处理后的原始文件_联通流量详单.xlsx'
        sorted_file_for_internal = PATH + '2_内部使用_排序后的_联通流量详单.xlsx'
        sorted_file_for_customer = PATH + '3_排序后的_联通流量详单.xlsx'  # 用于最后按照时间排序后的文件名字
        fileName = PATH + '联通流量详单.xls'                  #有些版本的Excel，不会显示后缀.xls  在这里人工添加上

        if not os.path.exists(fileName):
            print('输入的文件 %s 不存在，请检查后重新输入.' % fileName)
            exit()

        print('您输入的文件名后缀为.xls，将自动为您转换成.xlsx格式，开始处理...')
        fileName = xls2xlsx(fileName)
        cuc_milisecond_time_final(fileName)      # 调用函数进行时间格式转换
    else:
        raw_file_after_time_formated = PATH + '1_时间处理后的原始文件_' + fileName + '.xlsx'
        raw_file_after_time_formated_bk = PATH + '备份_1_时间处理后的原始文件_' + fileName + '.xlsx'
        sorted_file_for_internal = PATH + '2_内部使用_排序后的_' + fileName + '.xlsx'
        sorted_file_for_customer = PATH + '3_排序后的_' + fileName + '.xlsx'
        fileName = PATH + fileName + '.xls'

        if not os.path.exists(fileName):
            print('输入的文件 %s 不存在，请检查后重新输入.' % fileName)
            exit()
        if fileName.endswith('.xls'):
            #print('您输入的文件后缀为.xls，将自动为您转换成.xlsx格式，开始处理...')
            print('文件正在处理中......')
            fileName = xls2xlsx(fileName)
            data = pd.read_excel(fileName, engine='openpyxl')  # 用openpyxl代替xlrd打开.xlsx文件,否则会报错  #xiao
            data.drop(data.columns[10:11], inplace=True, axis=1)  # 删除列从列 0 开始数起，删除第10列, 第10列是邹波新添加进来的
            data.to_excel(fileName, index=False)
            cuc_milisecond_time_final(fileName)      # 调用函数进行时间格式转换
        elif fileName.endswith('.xlsx'):
            fileName = fileName
            data = pd.read_excel(fileName, engine='openpyxl')  # 用openpyxl代替xlrd打开.xlsx文件,否则会报错  #xiao
            data.drop(data.columns[10:11], inplace=True, axis=1)  # 删除列从列 0 开始数起，删除第10列 inplace=True 在当前基础上修改
            data.to_excel(fileName, index=False)
            cuc_milisecond_time_final(fileName)      # 调用函数进行时间格式转换
        else:
            print('输入的文件不符合excel格式，请检查后重新输入.')

    # 进行排序
    data = pd.read_excel(fileName, engine = 'openpyxl')  #用openpyxl代替xlrd打开.xlsx文件,否则会报错
    data.sort_values(by=['计费开始时间'], inplace= True, ascending=True)    #inplace=True 在当前内容中进行排序并保存

    data.to_excel(raw_file_after_time_formated, index=False)
    print('时间转后的原始数据文件保存在', raw_file_after_time_formated)

    # 复制一个raw_file_after_time_formated备份文件,用于处理成内部使用的简化版格式
    shutil.copy(raw_file_after_time_formated, raw_file_after_time_formated_bk)

    # 生成一个公司内部使用的简化版文件，保留一些必要的字段，供处理故障使用
    data1 = pd.read_excel(raw_file_after_time_formated_bk, engine='openpyxl')
    #data1.drop(data1.columns[10:11], inplace=True, axis=1) # 删除列从列 0 开始数起，删除第10列inplace=True 在当前基础上修改
    data1.drop(data1.columns[0:2], inplace=True, axis=1)  # 删除列从列 0 开始数起，inplace=True 在当前基础上修改
    data1.drop(data1.columns[5:7], inplace=True, axis=1)  # 删除之后，内存中重新排序，这行在新的基础上重新从 0 开始数起来做删除
    data1.drop(data1.columns[6:8], inplace=True, axis=1)
    data1.drop(data1.columns[11:13], inplace=True, axis=1)
    data1.drop(data1.columns[13:14], inplace=True, axis=1)
    # data1.drop(data1.columns[21:22], inplace=True, axis=1)
    data1.to_excel(sorted_file_for_internal, index=False)  # index=False 不要行索引， 保存文件
    print('供内部使用的简化版文件保存在', sorted_file_for_internal)

    delete_file(raw_file_after_time_formated_bk)  # 临时备份文件删除

    #生成可以提供给客户简化版格式的文件
    data2 = pd.read_excel(raw_file_after_time_formated, engine='openpyxl')
    data2.drop(data2.columns[0:3], inplace=True, axis=1)
    data2.drop(data2.columns[1:3], inplace=True, axis=1)
    data2.drop(data2.columns[2:9], inplace=True, axis=1)
    data2.drop(data2.columns[4:7], inplace=True, axis=1)
    data2.drop(data2.columns[6:9], inplace=True, axis=1)
    data2.to_excel(sorted_file_for_customer, index=False)         # index=False 不要行索引， 保存文件
    print('发给客户用的简化版文件保存在', sorted_file_for_customer)

    #提交给客户版本的文件中需要 剔除 tj.gxapn 话单
    data3 = pd.read_excel(sorted_file_for_customer, sheet_name='Sheet1')  # 读入需要处理的表格及sheet1页
    list = ['tj.gxapn', ]  # 定义需要处理掉的对象
    data3 = data3[-data3.apn.isin(list)]  # 删除 apn 列里含有tj.apn所在的行
    # #data1 = data[-data['转发'].str.contains('|'.join(list))]  # 然后再删除“转发”列里有这些ID的
    data3.to_excel(sorted_file_for_customer, index=False)  # 将处理后的结果写入新表

    #对 sorted_file_for_customer 文件里固定7列的列宽进行人工设置
    workbook = load_workbook(sorted_file_for_customer)
    sheet = workbook.active
    rows = sheet.max_row                  # ws.max_row 计算工作表行数，
    columns = sheet.max_column            # ws.max_column 计算作表列数

    data_sum = 0

    for row in range(2, sheet.max_row + 1):
        apn_name = str(sheet.cell(row, 1).value)     #第1列从1开始数起
        if apn_name == "ecarxinternet.gxapn":
            sheet.cell(row, 1).value = "公网"
        if apn_name == "zjyktkj.gxapn":
            sheet.cell(row, 1).value = "私网"

        data_volume = int(sheet.cell(row, 5).value) + int(sheet.cell(row, 6).value)  #上下行总流量
        sheet.cell(row, 7).value = data_volume
        data_sum += data_volume

    sheet.cell(1, 7).value = "流量Byte"
    sheet.cell(1, 8).value = "流量汇总Byte"
    sheet.cell(1, 9).value = "流量汇总MB"
    sheet.cell(2, 8).value = data_sum
    sheet.cell(2, 9).value = int(data_sum / 1024 / 1024)

    workbook.save(sorted_file_for_customer)  # 保存文件

    # 调用自定义函数来做居中处理
    excel_center(file_name = sorted_file_for_customer)
    data = pd.read_excel(sorted_file_for_customer, engine='openpyxl')
    data.drop(data.columns[4:6], inplace=True, axis=1)
    data.to_excel(sorted_file_for_customer, index=False)

    # 设定列宽度
    workbook = openpyxl.load_workbook(sorted_file_for_customer)
    sheet = workbook.active

    sheet.column_dimensions['A'].width = 8
    sheet.column_dimensions['B'].width = 15
    sheet.column_dimensions['C'].width = 25
    sheet.column_dimensions['D'].width = 25
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 20
    sheet.column_dimensions['G'].width = 20

    workbook.save(sorted_file_for_customer)

    # delete_file(fileName)   # 删除这个临时文件
# 处理联通短信sms文件
def cuc_sms_process():
    # 输入待处理的数据文件名（带完整后缀名称）
    PATH = 'C:\\pythonProject\\data\\'  # PATH 定义了数据文件的绝对路径
    fileName = input('请输入带后缀.xls或者.xlsx的完整文件名： 若直接回车，则默认输入文件名为 "联通短信详单.xls" : ')
    if fileName == "":
        fileName = PATH + '联通短信详单.xls'
        print(os.path.exists(fileName))
        if not os.path.exists(fileName):
            print('输入的文件不存在，请检查后重新输入.')
            exit()
        print('您输入的文件名后缀为.xls，将自动为您转换成.xlsx格式，开始处理...')
        fileName = xls2xlsx(fileName)

        cuc_yyyymmdd_time_final(fileName)          #调用函数进行时间格式转换
    else:
        fileName = PATH + fileName + '.xls'   #有些版本的Excel，不会显示后缀.xls  在这里人工添加上
        if not os.path.exists(fileName):
            print('输入的文件不存在，请检查后重新输入.')
            exit()
        if fileName.endswith('.xls'):
            #print('您输入的文件后缀为.xls，将自动为您转换成.xlsx格式，开始处理...')
            print('文件正在处理中......')
            fileName = xls2xlsx(fileName)
            cuc_yyyymmdd_time_final(fileName)      # 调用函数进行时间格式转换
        elif fileName.endswith('.xlsx'):
            fileName = fileName
            cuc_yyyymmdd_time_final(fileName)      # 调用函数进行时间格式转换
        else:
            print('输入的文件不符合excel格式，请检查后重新输入.')

# 处理移动流量data文件
def cmc_data_process():
    # 输入待处理的数据文件名（带完整后缀名称）
    PATH = 'C:\\pythonProject\\data\\'  # PATH 定义了数据文件的绝对路径
    fileName = input('请输入带后缀.xls或者.xlsx的完整文件名： 若直接回车，则默认输入文件名为 "移动流量详单.xls" : ')
    if fileName == "":
        fileName = PATH + '移动流量详单.xls'
        if not os.path.exists(fileName):
            print('输入的文件不存在，请检查后重新输入.')
            exit()
        print('您输入的文件名后缀为.xls，将自动为您转换成.xlsx格式，开始处理...')
        fileName = xls2xlsx(fileName)

        cmc_milisecond_time_final(fileName)      # 调用函数进行时间格式转换
    else:
        fileName = PATH + fileName + '.xls'   #有些版本的Excel，不会显示后缀.xls  在这里人工添加上
        if not os.path.exists(fileName):
            print('输入的文件不存在，请检查后重新输入.')
            exit()
        if fileName.endswith('.xls'):
            #print('您输入的文件后缀为.xls，将自动为您转换成.xlsx格式，开始处理...')
            print('文件正在处理中......')
            fileName = xls2xlsx(fileName)
            cmc_milisecond_time_final(fileName)      # 调用函数进行时间格式转换
        elif fileName.endswith('.xlsx'):
            fileName = fileName
            cmc_milisecond_time_final(fileName)      # 调用函数进行时间格式转换
        else:
            print('输入的文件不符合excel格式，请检查后重新输入.')

# 处理移动短信sms文件
def cmc_sms_process():
    # 输入待处理的数据文件名（带完整后缀名称）
    PATH = 'C:\\pythonProject\\data\\'  # PATH 定义了数据文件的绝对路径
    fileName = input('请输入带后缀.xls或者.xlsx的完整文件名： 若直接回车，则默认输入文件名为 "移动短信详单.xls" : ')
    if fileName == "":
        fileName = PATH + '移动短信详单.xls'
        if not os.path.exists(fileName):
            print('输入的文件不存在，请检查后重新输入.')
            exit()
        print('您输入的文件名后缀为.xls，将自动为您转换成.xlsx格式，开始处理...')
        fileName = xls2xlsx(fileName)
        cmc_yyyymmdd_time_final(fileName)      # 调用函数进行时间格式转换
    else:
        fileName = PATH + fileName + '.xls'   #有些版本的Excel，不会显示后缀.xls  在这里人工添加上
        if not os.path.exists(fileName):
            print('输入的文件不存在，请检查后重新输入.')
            exit()
        if fileName.endswith('.xls'):
            #print('您输入的文件后缀为.xls，将自动为您转换成.xlsx格式，开始处理...')
            print('文件正在处理中......')
            fileName = xls2xlsx(fileName)
            cmc_yyyymmdd_time_final(fileName)      # 调用函数进行时间格式转换
        elif fileName.endswith('.xlsx'):
            fileName = fileName
            cmc_yyyymmdd_time_final(fileName)      # 调用函数进行时间格式转换
        else:
            print('输入的文件不符合excel格式，请检查后重新输入.')

# 主程序，选择处理不同运营商的不同类型文件
def main():
    key = int(input('联通流量详单文件, 请按“1”，\n' 
                    '联通短信详单文件, 请按“2”，\n'
                    '移动流量详单文件, 请按“3”，\n'
                    '移动短信详单文件, 请按“4”，\n'
                    '其它按键则退出.')
              )
    print('======================================================')
    if key == 1:
        cuc_data_process()
    elif key == 2:
        cuc_sms_process()
    elif key == 3:
        cmc_data_process()
    elif key == 4:
        cmc_sms_process()
    else:
        print('您输入有误,请重新输入.')
        exit()

if __name__ == "__main__":
    main()

